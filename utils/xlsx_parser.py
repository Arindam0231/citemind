"""
XLSX Parser — extracts all sheets into structured dicts.
"""

import base64
import json
import io
from openpyxl import load_workbook
import pandas as pd
from utils.ingestion_service.data_service import DataIngestionService
from db.queries import get_excel_sheet_data


def parse_xlsx(contents_b64: str) -> dict:
    """
    Decode a base64-encoded XLSX file (from Dash Upload),
    read every sheet, and return a dict:
        {
            "Sheet1": [["Header1", "Header2"], ["val1", "val2"], ...],
            "Sheet2": [...],
        }
    Cell values are converted to strings. None cells become "".
    """
    if "," in contents_b64:
        contents_b64 = contents_b64.split(",", 1)[1]

    raw = base64.b64decode(contents_b64)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    sheets: dict[str, list[list[str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            str_row = [str(cell) if cell is not None else "" for cell in row]
            # Skip fully empty rows
            if any(c.strip() for c in str_row):
                rows.append(str_row)
        sheets[sheet_name] = rows

    wb.close()
    return sheets


def format_sheets_for_prompt(sheets: dict) -> str:
    ds = DataIngestionService()
    """
    Format parsed Excel sheets into a readable string for LLM context.
    Includes sheet name headers, column letters, and row numbers.
    """
    if not sheets:
        return "(No Excel data loaded)"
    complete_data = {}
    Context_Complete = {}
    for sheet_name, sheet_details in sheets.items():
        columns = {}
        for cell in sheet_details["cells"]:
            if not cell["is_header"]:
                columns[cell["col_index"]] = columns.get(cell["col_index"], [])
                columns[cell["col_index"]].append(cell["display_value"])
        for c_in, c_val in columns.items():
            print(c_in, len(c_val))
        for c_in, c_name in enumerate(sheet_details["headers"]):
            print(c_in, len(columns.get(c_in, [])), c_name)
        dataFrame = pd.DataFrame(
            {
                c_name: columns.get(c_in, [])
                for c_in, c_name in enumerate(sheet_details["headers"])
            }
        )

        complete_data[sheet_name] = dataFrame
        Context_Complete[sheet_name] = {
            # "data": dataFrame,
            "profile": ds.profile_dataframe(dataFrame),
            "column_types": ds.detect_column_types(dataFrame),
            # "categorical_insights": ds.gather_categorical_insights(
            #     dataFrame, ds.detect_column_types(dataFrame)
            # ),
            "llm_insights": get_excel_sheet_data(sheet_details["id"]).get(
                "llm_insights", {}
            ),
        }
    print("Context_Complete:", Context_Complete)
    return json.dumps(Context_Complete)


def _col_letter(index: int) -> str:
    """Convert 0-based column index to Excel-style letter (A, B, ... Z, AA, AB, ...)."""
    result = ""
    idx = index
    while True:
        result = chr(65 + idx % 26) + result
        idx = idx // 26 - 1
        if idx < 0:
            break
    return result
