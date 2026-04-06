"""
CiteMind — Excel parser with cell-level extraction.
Each cell stored with row_context for AI disambiguation.
"""

from __future__ import annotations

import base64
import datetime
import hashlib
import io
import json
import os
import re
import time
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from utils.ingestion_service.data_service import DataIngestionService
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def cleanup_processed_dir(processed_dir: str, days: int = 1):
    """Clean up files in processed_dir older than given days."""
    if not os.path.exists(processed_dir):
        return
    current_time = time.time()
    for filename in os.listdir(processed_dir):
        filepath = os.path.join(processed_dir, filename)
        if os.path.isfile(filepath):
            if current_time - os.path.getmtime(filepath) > (days * 86400):
                try:
                    os.remove(filepath)
                except Exception:
                    pass


def extract_numeric(value: Any) -> Optional[float]:
    """Extract a numeric value, stripping currency / commas / percent."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = re.sub(r"[₹$€£¥,% ]", "", value)
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None
    return None


def classify_type(cell: Any) -> str:
    """Classify an openpyxl cell into a data type string."""
    if cell.data_type == "n":
        return "number"
    elif cell.data_type == "s":
        return "string"
    elif cell.data_type == "b":
        return "boolean"
    elif cell.data_type == "d":
        return "date"
    elif cell.data_type == "f":
        return "formula"
    elif cell.data_type == "e":
        return "error"
    return "string"


def format_display(cell: Any) -> str:
    """Format a cell value for display."""
    if cell.value is None:
        return ""
    return str(cell.value)


def detect_headers(ws: Any) -> Dict[int, str]:
    """
    Auto-detect header row (assume row 1) and return
    {column_number: header_name} mapping.
    """
    headers = {}  # type: Dict[int, str]
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.column] = str(cell.value).strip()
    return headers


def parse_workbook(
    contents_b64: str,
    filename: Optional[str] = None,
) -> dict:
    """
    Parse a base64-encoded XLSX (from Dash Upload).

    Returns dict:
        {
            "sha256": str,
            "sheet_names": [str, ...],
            "sheets": {
                "SheetName": {
                    "sheet_index": int,
                    "row_count": int,
                    "col_count": int,
                    "header_row": int,
                    "headers": [str, ...],
                    "cells": [cell_dict, ...]
                }
            }
        }
    """
    # Strip data-URI prefix
    if "," in contents_b64:
        contents_b64 = contents_b64.split(",", 1)[1]

    raw = base64.b64decode(contents_b64)
    sha = hashlib.sha256(raw).hexdigest()
    wb = load_workbook(io.BytesIO(raw), data_only=True)

    log_start = datetime.datetime.now()
    data_ingestion_service = DataIngestionService()
    data_ingested = data_ingestion_service.register_data({filename: wb})
    print("Data ingestion time:", datetime.datetime.now() - log_start)
    ingested_wb = openpyxl.Workbook()
    ingested_wb.remove(ingested_wb.active)  # removes default "Sheet" created on init
    for sheet_info in data_ingested[filename]:
        ws = ingested_wb.create_sheet(title=sheet_info["sheet_name"])
        df = sheet_info["ProcessedDF"]
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)

        for r_idx, row in enumerate(
            df.itertuples(index=False), start=2
        ):  # start=2, header is row 1
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    def _extract_sheets_from_wb(target_wb) -> Dict[str, dict]:
        res_sheets = {}
        for sheet_idx, sheet_name in enumerate(target_wb.sheetnames):
            ws = target_wb[sheet_name]
            headers = detect_headers(ws)
            header_names = [headers.get(i + 1, "") for i in range(ws.max_column or 0)]

            cells = []  # type: List[dict]
            max_row = 0
            max_col = 0

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    raw_value = str(cell.value)
                    numeric = extract_numeric(cell.value)

                    # Build row context: {header: value} for this row
                    row_ctx = {}  # type: Dict[str, str]
                    for c in ws[cell.row]:
                        if c.column in headers and c.value is not None:
                            row_ctx[headers[c.column]] = str(c.value)

                    cells.append(
                        {
                            "cell_address": cell.coordinate,
                            "row_index": cell.row - 1,
                            "col_index": cell.column - 1,
                            "raw_value": raw_value,
                            "numeric_value": numeric,
                            "data_type": classify_type(cell),
                            "display_value": format_display(cell),
                            "row_context": json.dumps(row_ctx),
                            "is_header": cell.row == 1,
                        }
                    )

                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

            res_sheets[sheet_name] = {
                "sheet_index": sheet_idx,
                "row_count": max_row,
                "col_count": max_col,
                "header_row": 1 if headers else None,
                "headers": header_names,
                "cells": cells,
            }
        return res_sheets

    original_sheets = _extract_sheets_from_wb(wb)
    cleaned_sheets = _extract_sheets_from_wb(ingested_wb)
    # Add custom identifier to the workbook metadata
    ingested_wb.properties.identifier = f"citemind_{sha}"

    # Save the processed workbook
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    processed_dir = os.path.join(project_root, "processed")
    os.makedirs(processed_dir, exist_ok=True)
    cleanup_processed_dir(processed_dir)

    processed_path = os.path.join(processed_dir, f"{sha}_processed.xlsx")
    ingested_wb.save(processed_path)

    ingested_wb.close()
    wb.close()
    return {
        "sha256": sha,
        "sheet_names": list(ingested_wb.sheetnames),
        "original_sheets": original_sheets,
        "cleaned_sheets": cleaned_sheets,
        "processed_path": processed_path,
        "ingestion_report": {
            sheet_data["sheet_name"]: sheet_data.get("ReportTransformation", {})
            for sheet_data in data_ingested[filename]
        },
    }


def get_raw_bytes(contents_b64: str) -> bytes:
    """Decode base64 content to raw bytes."""
    if "," in contents_b64:
        contents_b64 = contents_b64.split(",", 1)[1]
    return base64.b64decode(contents_b64)


def format_sheets_for_prompt(sheets_data: dict) -> str:
    """Format parsed Excel data for LLM context."""
    if not sheets_data:
        return "(No Excel data loaded)"

    lines = []  # type: List[str]
    for sheet_name, info in sheets_data.items():
        lines.append('=== Sheet: "{}" ==='.format(sheet_name))
        headers = info.get("headers", [])
        if headers:
            hdr_parts = []  # type: List[str]
            for i, h in enumerate(headers):
                col_letter = get_column_letter(i + 1)
                hdr_parts.append("Col {}: {}".format(col_letter, h))
            lines.append("     " + " | ".join(hdr_parts))
            lines.append("     " + "-" * 60)

        cells = info.get("cells", [])
        # Group cells by row
        rows_map = {}  # type: Dict[int, List[dict]]
        for c in cells:
            ri = c["row_index"]
            if ri == 0:
                continue  # skip header row
            if ri not in rows_map:
                rows_map[ri] = []
            rows_map[ri].append(c)

        for ri in sorted(rows_map.keys()):
            row_cells = sorted(rows_map[ri], key=lambda x: x["col_index"])
            vals = [c.get("display_value", "") for c in row_cells]
            lines.append("Row {:>3}: {}".format(ri + 1, " | ".join(vals)))

        lines.append("")

    return "\n".join(lines)
